"""
Excel Export functionality for acoustic analysis results
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from models import get_session, Project, Space, Drawing
from models.hvac import HVACPath, HVACComponent, HVACSegment
from calculations import HVACPathCalculator, NCRatingAnalyzer


@dataclass
class ExportOptions:
    """Options for Excel export"""
    include_spaces: bool = True
    include_hvac_paths: bool = True
    include_components: bool = True
    include_rt60_details: bool = True
    include_nc_analysis: bool = True
    include_recommendations: bool = True
    include_charts: bool = True


class ExcelExporter:
    """Excel export functionality for acoustic analysis results"""
    
    def __init__(self):
        """Initialize the Excel exporter"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export functionality")
        
        self.hvac_calculator = HVACPathCalculator()
        self.nc_analyzer = NCRatingAnalyzer()
        
        # Styling
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def export_project_analysis(self, project_id: int, export_path: str, options: ExportOptions = None) -> bool:
        """
        Export complete project acoustic analysis to Excel
        
        Args:
            project_id: Project ID to export
            export_path: Path for Excel file output
            options: Export options
            
        Returns:
            True if export successful
        """
        if options is None:
            options = ExportOptions()
        
        try:
            # Get project data
            session = get_session()
            project = session.query(Project).filter(Project.id == project_id).first()
            
            if not project:
                session.close()
                raise ValueError(f"Project with ID {project_id} not found")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self.create_summary_sheet(wb, project, session)
            
            # Create sheets based on options
            if options.include_spaces:
                self.create_spaces_sheet(wb, project, session, options)
            
            if options.include_hvac_paths:
                self.create_hvac_paths_sheet(wb, project, session, options)
            
            if options.include_components:
                self.create_components_sheet(wb, project, session)
            
            if options.include_nc_analysis:
                self.create_nc_analysis_sheet(wb, project, session, options)
            
            # Save workbook
            wb.save(export_path)
            session.close()
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            if 'session' in locals():
                session.close()
            return False
    
    def create_summary_sheet(self, wb, project, session):
        """Create project summary sheet"""
        ws = wb.create_sheet("Project Summary", 0)
        
        # Project header
        ws['A1'] = "ACOUSTIC ANALYSIS REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Project info
        row = 3
        ws[f'A{row}'] = "Project Information"
        self.apply_header_style(ws, f'A{row}:F{row}')
        
        row += 1
        project_info = [
            ("Project Name", project.name),
            ("Description", project.description or ""),
            ("Created Date", project.created_date.strftime("%Y-%m-%d") if project.created_date else ""),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for label, value in project_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.subheader_font
            row += 1
        
        # Project statistics
        row += 1
        ws[f'A{row}'] = "Project Statistics"
        self.apply_header_style(ws, f'A{row}:F{row}')
        
        row += 1
        spaces = session.query(Space).filter(Space.project_id == project.id).all()
        hvac_paths = session.query(HVACPath).filter(HVACPath.project_id == project.id).all()
        hvac_components = session.query(HVACComponent).filter(HVACComponent.project_id == project.id).all()
        drawings = session.query(Drawing).filter(Drawing.project_id == project.id).all()
        
        stats = [
            ("Total Spaces", len(spaces)),
            ("Total HVAC Paths", len(hvac_paths)),
            ("Total HVAC Components", len(hvac_components)),
            ("Total Drawings", len(drawings)),
            ("Spaces with RT60 Calculated", len([s for s in spaces if s.calculated_rt60])),
            ("HVAC Paths with Noise Calculated", len([p for p in hvac_paths if p.calculated_noise])),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.subheader_font
            row += 1
        
        # Performance summary
        if spaces:
            row += 1
            ws[f'A{row}'] = "Performance Summary"
            self.apply_header_style(ws, f'A{row}:F{row}')
            
            row += 1
            # RT60 statistics
            rt60_values = [s.calculated_rt60 for s in spaces if s.calculated_rt60]
            if rt60_values:
                avg_rt60 = sum(rt60_values) / len(rt60_values)
                min_rt60 = min(rt60_values)
                max_rt60 = max(rt60_values)
                
                ws[f'A{row}'] = "Average RT60"
                ws[f'B{row}'] = f"{avg_rt60:.2f} seconds"
                ws[f'A{row}'].font = self.subheader_font
                row += 1
                
                ws[f'A{row}'] = "RT60 Range"
                ws[f'B{row}'] = f"{min_rt60:.2f} - {max_rt60:.2f} seconds"
                ws[f'A{row}'].font = self.subheader_font
                row += 1
            
            # NC rating statistics
            nc_values = [p.calculated_nc for p in hvac_paths if p.calculated_nc]
            if nc_values:
                avg_nc = sum(nc_values) / len(nc_values)
                min_nc = min(nc_values)
                max_nc = max(nc_values)
                
                ws[f'A{row}'] = "Average NC Rating"
                ws[f'B{row}'] = f"NC-{avg_nc:.0f}"
                ws[f'A{row}'].font = self.subheader_font
                row += 1
                
                ws[f'A{row}'] = "NC Range"
                ws[f'B{row}'] = f"NC-{min_nc} to NC-{max_nc}"
                ws[f'A{row}'].font = self.subheader_font
                row += 1
        
        # Auto-size columns
        self.auto_size_columns(ws)
    
    def create_spaces_sheet(self, wb, project, session, options):
        """Create spaces analysis sheet"""
        ws = wb.create_sheet("Spaces Analysis")
        
        spaces = session.query(Space).filter(Space.project_id == project.id).all()
        
        if not spaces:
            ws['A1'] = "No spaces found in project"
            return
        
        # Headers
        headers = [
            "Space Name", "Description", "Floor Area (sf)", "Ceiling Height (ft)",
            "Volume (cf)", "Wall Area (sf)", "Target RT60 (s)", "Calculated RT60 (s)",
            "Ceiling Material", "Wall Material", "Floor Material"
        ]
        
        row = 1
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self.apply_header_style(ws, f'{get_column_letter(col)}{row}')
        
        # Data rows
        for space in spaces:
            row += 1
            data = [
                space.name,
                space.description or "",
                space.floor_area or 0,
                space.ceiling_height or 0,
                space.volume or 0,
                space.wall_area or 0,
                space.target_rt60 or 0,
                space.calculated_rt60 or 0,
                space.ceiling_material or "",
                space.wall_material or "",
                space.floor_material or ""
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
                if col > 2 and isinstance(value, (int, float)):
                    ws.cell(row=row, column=col).number_format = '0.00'
        
        # RT60 compliance analysis
        if options.include_rt60_details:
            row += 3
            ws[f'A{row}'] = "RT60 Compliance Analysis"
            self.apply_header_style(ws, f'A{row}:D{row}')
            
            row += 1
            compliance_headers = ["Space Name", "Target RT60", "Calculated RT60", "Status"]
            for col, header in enumerate(compliance_headers, 1):
                ws.cell(row=row, column=col, value=header)
                self.apply_subheader_style(ws, f'{get_column_letter(col)}{row}')
            
            for space in spaces:
                if space.calculated_rt60 and space.target_rt60:
                    row += 1
                    difference = abs(space.calculated_rt60 - space.target_rt60)
                    tolerance = space.target_rt60 * 0.1  # 10% tolerance
                    
                    status = "Within Tolerance" if difference <= tolerance else "Outside Tolerance"
                    
                    data = [space.name, space.target_rt60, space.calculated_rt60, status]
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if col == 4 and status == "Outside Tolerance":
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        self.auto_size_columns(ws)
    
    def create_hvac_paths_sheet(self, wb, project, session, options):
        """Create HVAC paths analysis sheet"""
        ws = wb.create_sheet("HVAC Paths")
        
        hvac_paths = session.query(HVACPath).filter(HVACPath.project_id == project.id).all()
        
        if not hvac_paths:
            ws['A1'] = "No HVAC paths found in project"
            return
        
        # Headers
        headers = [
            "Path Name", "Path Type", "Description", "Target Space",
            "Calculated Noise (dB)", "NC Rating", "Segment Count",
            "Total Length (ft)", "Created Date"
        ]
        
        row = 1
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self.apply_header_style(ws, f'{get_column_letter(col)}{row}')
        
        # Data rows
        for path in hvac_paths:
            row += 1
            total_length = sum(seg.length or 0 for seg in path.segments)
            
            data = [
                path.name,
                path.path_type or "supply",
                path.description or "",
                path.target_space.name if path.target_space else "",
                path.calculated_noise or 0,
                path.calculated_nc or 0,
                len(path.segments),
                total_length,
                path.created_date.strftime("%Y-%m-%d") if path.created_date else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col in [5, 6, 7, 8] and isinstance(value, (int, float)):
                    cell.number_format = '0.00' if col == 8 else '0'
                
                # Highlight high NC ratings
                if col == 6 and isinstance(value, (int, float)) and value > 45:
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Detailed segment analysis
        if options.include_nc_analysis:
            row += 3
            ws[f'A{row}'] = "Detailed Path Analysis"
            self.apply_header_style(ws, f'A{row}:I{row}')
            
            row += 1
            segment_headers = [
                "Path Name", "Segment #", "Length (ft)", "Duct Size",
                "Duct Type", "Distance Loss (dB)", "Duct Loss (dB)",
                "Fitting Additions (dB)", "Noise After (dB)"
            ]
            
            for col, header in enumerate(segment_headers, 1):
                ws.cell(row=row, column=col, value=header)
                self.apply_subheader_style(ws, f'{get_column_letter(col)}{row}')
            
            # Calculate detailed results for each path
            for path in hvac_paths:
                if path.segments:
                    result = self.hvac_calculator.calculate_path_noise(path.id)
                    
                    for i, segment_result in enumerate(result.segment_results):
                        row += 1
                        segment = path.segments[i] if i < len(path.segments) else None
                        
                        duct_size = ""
                        if segment and segment.duct_width and segment.duct_height:
                            duct_size = f"{segment.duct_width}x{segment.duct_height}"
                        
                        data = [
                            path.name,
                            segment_result.get('segment_number', i+1),
                            segment_result.get('length', 0),
                            duct_size,
                            segment.duct_type if segment else "sheet_metal",
                            segment_result.get('distance_loss', 0),
                            segment_result.get('duct_loss', 0),
                            segment_result.get('fitting_additions', 0),
                            segment_result.get('noise_after', 0)
                        ]
                        
                        for col, value in enumerate(data, 1):
                            cell = ws.cell(row=row, column=col, value=value)
                            if col >= 3 and isinstance(value, (int, float)):
                                cell.number_format = '0.00'
        
        self.auto_size_columns(ws)
    
    def create_components_sheet(self, wb, project, session):
        """Create HVAC components sheet"""
        ws = wb.create_sheet("HVAC Components")
        
        components = session.query(HVACComponent).filter(HVACComponent.project_id == project.id).all()
        
        if not components:
            ws['A1'] = "No HVAC components found in project"
            return
        
        # Headers
        headers = [
            "Component Name", "Type", "Drawing", "X Position", "Y Position",
            "Noise Level (dB)", "Created Date"
        ]
        
        row = 1
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self.apply_header_style(ws, f'{get_column_letter(col)}{row}')
        
        # Data rows
        for component in components:
            row += 1
            data = [
                component.name,
                component.component_type,
                component.drawing.name if component.drawing else "",
                component.x_position or 0,
                component.y_position or 0,
                component.noise_level or 0,
                component.created_date.strftime("%Y-%m-%d") if component.created_date else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col in [4, 5, 6] and isinstance(value, (int, float)):
                    cell.number_format = '0.00'
        
        self.auto_size_columns(ws)
    
    def create_nc_analysis_sheet(self, wb, project, session, options):
        """Create NC analysis and recommendations sheet"""
        ws = wb.create_sheet("NC Analysis")
        
        hvac_paths = session.query(HVACPath).filter(HVACPath.project_id == project.id).all()
        spaces = session.query(Space).filter(Space.project_id == project.id).all()
        
        if not hvac_paths and not spaces:
            ws['A1'] = "No data available for NC analysis"
            return
        
        row = 1
        ws[f'A{row}'] = "NC Rating Analysis Summary"
        self.apply_header_style(ws, f'A{row}:F{row}')
        
        # NC distribution
        row += 2
        nc_headers = ["NC Rating", "Number of Paths", "Percentage", "Typical Applications"]
        for col, header in enumerate(nc_headers, 1):
            ws.cell(row=row, column=col, value=header)
            self.apply_subheader_style(ws, f'{get_column_letter(col)}{row}')
        
        # Count NC ratings
        nc_counts = {}
        total_paths = 0
        for path in hvac_paths:
            if path.calculated_nc:
                nc_rating = int(path.calculated_nc)
                nc_counts[nc_rating] = nc_counts.get(nc_rating, 0) + 1
                total_paths += 1
        
        # NC descriptions
        nc_descriptions = {
            20: "Executive offices, conference rooms",
            25: "Open offices, classrooms",
            30: "General offices, retail spaces",
            35: "Restaurants, lobbies",
            40: "Cafeterias, gymnasiums",
            45: "Workshops, mechanical rooms",
            50: "Industrial spaces"
        }
        
        for nc_rating in sorted(nc_counts.keys()):
            row += 1
            count = nc_counts[nc_rating]
            percentage = (count / total_paths) * 100 if total_paths > 0 else 0
            description = nc_descriptions.get(nc_rating, "Unknown application")
            
            data = [f"NC-{nc_rating}", count, f"{percentage:.1f}%", description]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Recommendations section
        if options.include_recommendations:
            row += 3
            ws[f'A{row}'] = "Recommendations"
            self.apply_header_style(ws, f'A{row}:F{row}')
            
            row += 1
            rec_headers = ["Path/Space", "Current NC", "Issue", "Recommendation"]
            for col, header in enumerate(rec_headers, 1):
                ws.cell(row=row, column=col, value=header)
                self.apply_subheader_style(ws, f'{get_column_letter(col)}{row}')
            
            # Generate recommendations for high NC paths
            for path in hvac_paths:
                if path.calculated_nc and path.calculated_nc > 35:
                    row += 1
                    issue = f"NC-{path.calculated_nc} exceeds typical comfort levels"
                    
                    if path.calculated_nc > 45:
                        recommendation = "Major noise control required - consider equipment replacement and duct silencing"
                    elif path.calculated_nc > 40:
                        recommendation = "Add duct silencers and improve equipment isolation"
                    else:
                        recommendation = "Consider acoustic lining and flexible connections"
                    
                    data = [path.name, f"NC-{path.calculated_nc}", issue, recommendation]
                    for col, value in enumerate(data, 1):
                        ws.cell(row=row, column=col, value=value)
        
        self.auto_size_columns(ws)
    
    def apply_header_style(self, ws, range_str):
        """Apply header styling to a range"""
        for row in ws[range_str]:
            for cell in row:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
    
    def apply_subheader_style(self, ws, range_str):
        """Apply subheader styling to a range"""
        for row in ws[range_str]:
            for cell in row:
                cell.font = self.subheader_font
                cell.fill = self.subheader_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
    
    def auto_size_columns(self, ws):
        """Auto-size all columns in worksheet"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_export_summary(self, project_id: int) -> Dict[str, Any]:
        """
        Get summary of what would be exported
        
        Args:
            project_id: Project ID
            
        Returns:
            Dictionary with export summary
        """
        try:
            session = get_session()
            project = session.query(Project).filter(Project.id == project_id).first()
            
            if not project:
                session.close()
                return {"error": "Project not found"}
            
            spaces = session.query(Space).filter(Space.project_id == project.id).all()
            hvac_paths = session.query(HVACPath).filter(HVACPath.project_id == project.id).all()
            hvac_components = session.query(HVACComponent).filter(HVACComponent.project_id == project.id).all()
            
            summary = {
                "project_name": project.name,
                "total_spaces": len(spaces),
                "spaces_with_rt60": len([s for s in spaces if s.calculated_rt60]),
                "total_hvac_paths": len(hvac_paths),
                "paths_with_noise": len([p for p in hvac_paths if p.calculated_noise]),
                "total_components": len(hvac_components),
                "sheets_to_export": [
                    "Project Summary",
                    "Spaces Analysis" if spaces else None,
                    "HVAC Paths" if hvac_paths else None,
                    "HVAC Components" if hvac_components else None,
                    "NC Analysis" if hvac_paths else None
                ]
            }
            
            # Filter out None sheets
            summary["sheets_to_export"] = [s for s in summary["sheets_to_export"] if s]
            
            session.close()
            return summary
            
        except Exception as e:
            return {"error": str(e)}

# Export the availability flag for compatibility
EXCEL_EXPORT_AVAILABLE = EXCEL_AVAILABLE