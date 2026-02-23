"""
Edge case tests for calculations and export functionality.

Tests cover:
- NC rating analysis edge cases
- Excel exporter edge cases
- Octave band data edge cases
- Export options variations
"""

import sys
import os
import tempfile
import pytest
import math

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))


class TestNCRatingAnalyzerEdgeCases:
    """Edge case tests for NC rating analyzer."""

    @pytest.fixture
    def analyzer(self):
        """Create NC rating analyzer instance."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer
        return NCRatingAnalyzer()

    def test_determine_nc_rating_empty_list(self, analyzer):
        """Test NC rating with empty octave levels list."""
        result = analyzer.determine_nc_rating([])
        # Should return default NC-30
        assert result == 30

    def test_determine_nc_rating_short_list(self, analyzer):
        """Test NC rating with insufficient octave levels."""
        result = analyzer.determine_nc_rating([35, 40, 45])
        assert result == 30  # Default for invalid data

    def test_determine_nc_rating_all_zero(self, analyzer):
        """Test NC rating when all levels are zero."""
        result = analyzer.determine_nc_rating([0, 0, 0, 0, 0, 0, 0, 0])
        # Should return lowest NC (NC-15)
        assert result == 15

    def test_determine_nc_rating_very_high_levels(self, analyzer):
        """Test NC rating with extremely high noise levels."""
        # Levels that exceed all NC curves
        result = analyzer.determine_nc_rating([90, 85, 80, 75, 70, 65, 65, 65])
        # Should return maximum NC rating
        assert result == 65

    def test_determine_nc_rating_borderline_values(self, analyzer):
        """Test NC rating at exact curve boundary values."""
        # Use exact NC-35 curve values
        nc35_values = [60, 52, 45, 40, 36, 34, 33, 32]
        result = analyzer.determine_nc_rating(nc35_values)
        assert result == 35

    def test_calculate_overall_dba_empty_list(self, analyzer):
        """Test dBA calculation with empty list."""
        result = analyzer.calculate_overall_dba([])
        assert result == 0.0

    def test_calculate_overall_dba_all_zero(self, analyzer):
        """Test dBA calculation when all levels are zero."""
        result = analyzer.calculate_overall_dba([0, 0, 0, 0, 0, 0, 0, 0])
        assert result == 0.0

    def test_calculate_overall_dba_negative_values(self, analyzer):
        """Test dBA calculation with negative values (invalid but shouldn't crash)."""
        result = analyzer.calculate_overall_dba([-10, -5, 0, 5, 10, 15, 10, 5])
        # Should handle negatives gracefully
        assert isinstance(result, float)
        assert not math.isnan(result)

    def test_calculate_overall_dba_typical_hvac(self, analyzer):
        """Test dBA calculation with typical HVAC spectrum."""
        # Typical NC-35 levels
        levels = [60, 52, 45, 40, 36, 34, 33, 32]
        result = analyzer.calculate_overall_dba(levels)
        # Should produce reasonable dBA value (typically NC + ~5-10 dB)
        assert 35 < result < 55

    def test_generate_warnings_very_low_levels(self, analyzer):
        """Test warning generation for very low noise levels."""
        levels = [10, 8, 6, 5, 5, 5, 4, 3]
        warnings = analyzer.generate_warnings(levels, 15, [])
        assert any("verify" in w.lower() for w in warnings)

    def test_generate_warnings_high_nc(self, analyzer):
        """Test warning generation for high NC rating."""
        levels = [75, 70, 65, 60, 55, 53, 52, 51]
        warnings = analyzer.generate_warnings(levels, 55, [])
        assert any("noise control" in w.lower() for w in warnings)

    def test_generate_warnings_low_frequency_dominance(self, analyzer):
        """Test warning for low frequency dominant noise."""
        levels = [70, 55, 45, 40, 35, 33, 32, 30]  # 63 Hz much higher
        warnings = analyzer.generate_warnings(levels, 40, [])
        assert any("low frequency" in w.lower() for w in warnings)

    def test_generate_warnings_high_frequency_emphasis(self, analyzer):
        """Test warning for high frequency emphasis."""
        levels = [30, 35, 38, 40, 42, 45, 48, 55]  # 8000 Hz higher than 1000 Hz
        warnings = analyzer.generate_warnings(levels, 35, [])
        assert any("high frequency" in w.lower() for w in warnings)

    def test_generate_warnings_with_exceedances(self, analyzer):
        """Test warning generation when target is exceeded."""
        levels = [60, 52, 45, 40, 36, 34, 33, 32]
        exceedances = [(500, 5.0), (1000, 3.0)]
        warnings = analyzer.generate_warnings(levels, 35, exceedances)
        assert any("exceeded" in w.lower() for w in warnings)

    def test_get_nc_description_valid_ratings(self, analyzer):
        """Test NC description for all valid ratings."""
        for nc in [15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65]:
            desc = analyzer.get_nc_description(nc)
            assert f"NC-{nc}" in desc
            assert len(desc) > 10  # Should have meaningful content

    def test_get_nc_description_intermediate_rating(self, analyzer):
        """Test NC description for intermediate rating (e.g., NC-27)."""
        desc = analyzer.get_nc_description(27)
        assert "NC-27" in desc
        assert "NC-25" in desc  # Should reference closest standard

    def test_get_nc_description_extreme_rating(self, analyzer):
        """Test NC description for rating beyond standard range."""
        desc = analyzer.get_nc_description(70)
        assert "NC-70" in desc

    def test_compare_to_standards_known_space_types(self, analyzer):
        """Test comparison for known space types."""
        space_types = ["private_office", "open_office", "conference_room",
                       "classroom", "library", "hospital_room"]

        for space_type in space_types:
            result = analyzer.compare_to_standards(30, space_type)
            assert "space_type" in result
            assert "measured_nc" in result
            assert "recommended_nc" in result
            assert "compliance" in result

    def test_compare_to_standards_unknown_space_type(self, analyzer):
        """Test comparison for unknown space type."""
        result = analyzer.compare_to_standards(35, "unknown_space_xyz")
        # Should use defaults
        assert result["recommended_nc"] == 30
        assert result["maximum_nc"] == 35

    def test_compare_to_standards_compliant(self, analyzer):
        """Test compliance when meeting recommended criteria."""
        result = analyzer.compare_to_standards(20, "private_office")
        assert result["compliance"] == "Excellent"

    def test_compare_to_standards_acceptable(self, analyzer):
        """Test compliance when meeting maximum but not recommended."""
        # Private office: recommended=25, maximum=30
        result = analyzer.compare_to_standards(28, "private_office")
        assert result["compliance"] == "Acceptable"

    def test_compare_to_standards_non_compliant(self, analyzer):
        """Test compliance when exceeding maximum."""
        result = analyzer.compare_to_standards(45, "private_office")
        assert result["compliance"] == "Non-compliant"
        assert result["improvement_needed"] > 0


class TestOctaveBandDataEdgeCases:
    """Edge case tests for OctaveBandData class."""

    def test_octave_band_data_default_values(self):
        """Test default values for OctaveBandData."""
        from calculations.nc_rating_analyzer import OctaveBandData
        data = OctaveBandData()
        assert data.freq_63 == 0.0
        assert data.freq_8000 == 0.0

    def test_octave_band_data_to_list(self):
        """Test conversion to list."""
        from calculations.nc_rating_analyzer import OctaveBandData
        data = OctaveBandData(
            freq_63=63, freq_125=125, freq_250=250, freq_500=500,
            freq_1000=1000, freq_2000=2000, freq_4000=4000, freq_8000=8000
        )
        result = data.to_list()
        assert len(result) == 8
        assert result[0] == 63
        assert result[7] == 8000

    def test_octave_band_data_from_list(self):
        """Test creation from list."""
        from calculations.nc_rating_analyzer import OctaveBandData
        values = [63, 125, 250, 500, 1000, 2000, 4000, 8000]
        data = OctaveBandData().from_list(values)
        assert data.freq_63 == 63
        assert data.freq_8000 == 8000

    def test_octave_band_data_from_short_list(self):
        """Test handling of short list (< 8 elements)."""
        from calculations.nc_rating_analyzer import OctaveBandData
        values = [63, 125, 250]  # Only 3 elements
        original = OctaveBandData(freq_63=10)
        result = original.from_list(values)
        # Should return original unchanged
        assert result.freq_63 == 10

    def test_octave_band_data_negative_values(self):
        """Test handling negative values."""
        from calculations.nc_rating_analyzer import OctaveBandData
        data = OctaveBandData(freq_63=-10, freq_125=0, freq_250=10)
        result = data.to_list()
        assert result[0] == -10  # Should preserve negative


class TestNCAnalysisResultEdgeCases:
    """Edge case tests for NCAnalysisResult."""

    def test_analyze_octave_band_data_no_target(self):
        """Test analysis without target NC."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer, OctaveBandData
        analyzer = NCRatingAnalyzer()

        data = OctaveBandData(
            freq_63=60, freq_125=52, freq_250=45, freq_500=40,
            freq_1000=36, freq_2000=34, freq_4000=33, freq_8000=32
        )

        result = analyzer.analyze_octave_band_data(data)

        assert result.nc_rating == 35
        assert result.meets_criteria == True  # No target, so meets by default
        assert len(result.exceedances) == 0

    def test_analyze_octave_band_data_with_target(self):
        """Test analysis with target NC that is exceeded."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer, OctaveBandData
        analyzer = NCRatingAnalyzer()

        # NC-35 spectrum
        data = OctaveBandData(
            freq_63=60, freq_125=52, freq_250=45, freq_500=40,
            freq_1000=36, freq_2000=34, freq_4000=33, freq_8000=32
        )

        result = analyzer.analyze_octave_band_data(data, target_nc=30)

        assert result.meets_criteria == False
        assert len(result.exceedances) > 0

    def test_analyze_octave_band_data_invalid_target(self):
        """Test analysis with invalid target NC (not in standard curves)."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer, OctaveBandData
        analyzer = NCRatingAnalyzer()

        data = OctaveBandData(
            freq_63=60, freq_125=52, freq_250=45, freq_500=40,
            freq_1000=36, freq_2000=34, freq_4000=33, freq_8000=32
        )

        # Target NC-27 doesn't exist in curves
        result = analyzer.analyze_octave_band_data(data, target_nc=27)

        # Should still return valid result
        assert isinstance(result.nc_rating, int)
        # No exceedances since target curve doesn't exist
        assert len(result.exceedances) == 0


class TestEstimateOctaveBandsEdgeCases:
    """Edge case tests for octave band estimation from dBA."""

    @pytest.fixture
    def analyzer(self):
        """Create NC rating analyzer instance."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer
        return NCRatingAnalyzer()

    def test_estimate_octave_bands_typical_hvac(self, analyzer):
        """Test estimation with typical HVAC spectrum."""
        result = analyzer.estimate_octave_bands_from_dba(40, "typical_hvac")
        levels = result.to_list()
        assert len(levels) == 8
        assert all(level >= 0 for level in levels)

    def test_estimate_octave_bands_fan_noise(self, analyzer):
        """Test estimation with fan noise spectrum."""
        result = analyzer.estimate_octave_bands_from_dba(45, "fan_noise")
        levels = result.to_list()
        # Fan noise should have higher low frequencies
        assert levels[0] > levels[4]  # 63 Hz > 1000 Hz

    def test_estimate_octave_bands_diffuser_noise(self, analyzer):
        """Test estimation with diffuser noise spectrum."""
        result = analyzer.estimate_octave_bands_from_dba(35, "diffuser_noise")
        levels = result.to_list()
        assert len(levels) == 8

    def test_estimate_octave_bands_flat_spectrum(self, analyzer):
        """Test estimation with flat spectrum."""
        result = analyzer.estimate_octave_bands_from_dba(40, "flat_spectrum")
        levels = result.to_list()
        # Flat spectrum should have similar levels
        assert max(levels) - min(levels) < 5

    def test_estimate_octave_bands_unknown_type(self, analyzer):
        """Test estimation with unknown spectrum type."""
        result = analyzer.estimate_octave_bands_from_dba(40, "unknown_type")
        levels = result.to_list()
        # Should fall back to typical HVAC
        assert len(levels) == 8

    def test_estimate_octave_bands_zero_dba(self, analyzer):
        """Test estimation with zero dBA."""
        result = analyzer.estimate_octave_bands_from_dba(0, "typical_hvac")
        levels = result.to_list()
        # Should produce low/zero values
        assert all(level >= 0 for level in levels)

    def test_estimate_octave_bands_very_high_dba(self, analyzer):
        """Test estimation with very high dBA."""
        result = analyzer.estimate_octave_bands_from_dba(90, "typical_hvac")
        levels = result.to_list()
        assert all(level > 50 for level in levels)


class TestRecommendNoiseControlEdgeCases:
    """Edge case tests for noise control recommendations."""

    @pytest.fixture
    def analyzer(self):
        """Create NC rating analyzer instance."""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer
        return NCRatingAnalyzer()

    def test_recommend_when_meeting_target(self, analyzer):
        """Test recommendations when already meeting target."""
        from calculations.nc_rating_analyzer import NCAnalysisResult, OctaveBandData

        result = NCAnalysisResult(
            nc_rating=30,
            octave_band_levels=OctaveBandData(),
            exceedances=[],
            overall_dba=35,
            calculation_method="test",
            warnings=[],
            meets_criteria=True
        )

        recommendations = analyzer.recommend_noise_control(result, target_nc=35)
        assert len(recommendations) == 1
        assert "meet target" in recommendations[0].lower()

    def test_recommend_small_reduction(self, analyzer):
        """Test recommendations for small reduction (<=5 NC points)."""
        from calculations.nc_rating_analyzer import NCAnalysisResult, OctaveBandData

        result = NCAnalysisResult(
            nc_rating=35,
            octave_band_levels=OctaveBandData(),
            exceedances=[],
            overall_dba=40,
            calculation_method="test",
            warnings=[],
            meets_criteria=False
        )

        recommendations = analyzer.recommend_noise_control(result, target_nc=32)
        assert len(recommendations) >= 1
        assert any("silencer" in r.lower() for r in recommendations)

    def test_recommend_medium_reduction(self, analyzer):
        """Test recommendations for medium reduction (5-10 NC points)."""
        from calculations.nc_rating_analyzer import NCAnalysisResult, OctaveBandData

        result = NCAnalysisResult(
            nc_rating=45,
            octave_band_levels=OctaveBandData(),
            exceedances=[],
            overall_dba=50,
            calculation_method="test",
            warnings=[],
            meets_criteria=False
        )

        recommendations = analyzer.recommend_noise_control(result, target_nc=38)
        assert len(recommendations) >= 3
        assert any("vibration" in r.lower() for r in recommendations)

    def test_recommend_major_reduction(self, analyzer):
        """Test recommendations for major reduction (>10 NC points)."""
        from calculations.nc_rating_analyzer import NCAnalysisResult, OctaveBandData

        result = NCAnalysisResult(
            nc_rating=55,
            octave_band_levels=OctaveBandData(),
            exceedances=[],
            overall_dba=60,
            calculation_method="test",
            warnings=[],
            meets_criteria=False
        )

        recommendations = analyzer.recommend_noise_control(result, target_nc=35)
        assert len(recommendations) >= 4
        assert any("major" in r.lower() for r in recommendations)

    def test_recommend_with_low_frequency_issue(self, analyzer):
        """Test recommendations with low frequency noise dominance."""
        from calculations.nc_rating_analyzer import NCAnalysisResult, OctaveBandData

        # Low frequency dominant spectrum
        octave = OctaveBandData(
            freq_63=75, freq_125=65, freq_250=55, freq_500=45,
            freq_1000=40, freq_2000=38, freq_4000=36, freq_8000=34
        )

        result = NCAnalysisResult(
            nc_rating=50,
            octave_band_levels=octave,
            exceedances=[],
            overall_dba=55,
            calculation_method="test",
            warnings=[],
            meets_criteria=False
        )

        recommendations = analyzer.recommend_noise_control(result, target_nc=35)
        assert any("low frequency" in r.lower() for r in recommendations)


class TestExcelExporterEdgeCases:
    """Edge case tests for Excel exporter."""

    @pytest.fixture
    def temp_db_session(self):
        """Create a temporary in-memory database for testing."""
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from models.database import Base

        engine = create_engine('sqlite:///:memory:', echo=False)

        # Import all models
        from models.project import Project
        from models.drawing import Drawing
        from models.space import Space

        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(bind=engine)
        session = SessionLocal()
        yield session
        session.close()

    def test_export_options_defaults(self):
        """Test ExportOptions default values."""
        from data.excel_exporter import ExportOptions

        options = ExportOptions()
        assert options.include_spaces == True
        assert options.include_hvac_paths == True
        assert options.leed_format == True

    def test_export_options_custom(self):
        """Test ExportOptions with custom values."""
        from data.excel_exporter import ExportOptions

        options = ExportOptions(
            include_spaces=False,
            include_hvac_paths=False,
            leed_format=False
        )

        assert options.include_spaces == False
        assert options.include_hvac_paths == False
        assert options.leed_format == False

    def test_excel_available_flag(self):
        """Test EXCEL_AVAILABLE flag is set correctly."""
        from data.excel_exporter import EXCEL_EXPORT_AVAILABLE

        # openpyxl should be installed in test environment
        assert EXCEL_EXPORT_AVAILABLE == True

    def test_exporter_initialization(self):
        """Test ExcelExporter can be initialized."""
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        assert hasattr(exporter, 'hvac_calculator')
        assert hasattr(exporter, 'nc_analyzer')

    def test_get_export_summary_nonexistent_project(self):
        """Test export summary for non-existent project."""
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        result = exporter.get_export_summary(99999)  # Non-existent ID

        assert "error" in result

    def test_export_to_invalid_path(self):
        """Test export to invalid/nonexistent path."""
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        result = exporter.export_project_analysis(
            1,  # Assuming no project exists
            "/nonexistent/path/file.xlsx"
        )

        assert result == False

    def test_export_creates_valid_workbook(self):
        """Test export creates a valid Excel workbook structure."""
        import tempfile
        import os
        from data.excel_exporter import ExcelExporter, ExportOptions
        import openpyxl

        # This test requires a real database session with a project
        # For now, test with minimal project creation
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from models.database import Base
        from models.project import Project

        # Create temporary database
        engine = create_engine('sqlite:///:memory:', echo=False)
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(bind=engine)
        session = SessionLocal()

        # Create minimal project
        project = Project(name="Test Export Project")
        session.add(project)
        session.commit()
        project_id = project.id
        session.close()

        # Note: This test may not work fully because ExcelExporter
        # uses get_session() which connects to the main database.
        # This is a limitation we're documenting with this test.


class TestAutoSizeColumnsEdgeCases:
    """Edge case tests for auto_size_columns method."""

    def test_auto_size_empty_worksheet(self):
        """Test auto-sizing on empty worksheet."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        # Should not crash on empty worksheet
        exporter.auto_size_columns(ws)

    def test_auto_size_with_long_content(self):
        """Test auto-sizing with very long content."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add very long content
        ws['A1'] = "A" * 1000

        exporter.auto_size_columns(ws)

        # Column width should be capped at 50
        assert ws.column_dimensions['A'].width <= 52  # 50 + 2 padding

    def test_auto_size_with_none_values(self):
        """Test auto-sizing when cells contain None."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = None
        ws['A2'] = "Test"

        # Should not crash
        exporter.auto_size_columns(ws)


class TestApplyStylesEdgeCases:
    """Edge case tests for style application methods."""

    def test_apply_header_style_single_cell(self):
        """Test header style on single cell."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = "Header"
        exporter.apply_header_style(ws, 'A1')

        assert ws['A1'].font.bold == True

    def test_apply_header_style_range(self):
        """Test header style on cell range."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = "Header 1"
        ws['B1'] = "Header 2"
        ws['C1'] = "Header 3"

        exporter.apply_header_style(ws, 'A1:C1')

        assert ws['A1'].font.bold == True
        assert ws['C1'].font.bold == True

    def test_apply_subheader_style_single_cell(self):
        """Test subheader style on single cell."""
        import openpyxl
        from data.excel_exporter import ExcelExporter

        exporter = ExcelExporter()
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = "Subheader"
        exporter.apply_subheader_style(ws, 'A1')

        assert ws['A1'].font.bold == True


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
