"""
Performance Profiling Tests for Acoustic Analysis Tool

This module provides comprehensive performance benchmarks for the calculation engines
and identifies potential bottlenecks for optimization.

Run with: pytest tests/test_performance_profiling.py -v -s
"""

import time
import cProfile
import pstats
import io
import sys
import os
from typing import Dict, List, Callable, Any
from dataclasses import dataclass
from statistics import mean, stdev

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import pytest


@dataclass
class BenchmarkResult:
    """Result of a benchmark run"""
    name: str
    iterations: int
    total_time: float
    avg_time: float
    min_time: float
    max_time: float
    std_dev: float
    ops_per_second: float

    def __str__(self):
        return (
            f"{self.name}:\n"
            f"  Iterations: {self.iterations}\n"
            f"  Total time: {self.total_time:.4f}s\n"
            f"  Average: {self.avg_time*1000:.3f}ms\n"
            f"  Min: {self.min_time*1000:.3f}ms\n"
            f"  Max: {self.max_time*1000:.3f}ms\n"
            f"  Std Dev: {self.std_dev*1000:.3f}ms\n"
            f"  Ops/sec: {self.ops_per_second:.1f}"
        )


def benchmark(func: Callable, iterations: int = 100, warmup: int = 5) -> BenchmarkResult:
    """Run a benchmark on a function"""
    # Warmup runs
    for _ in range(warmup):
        func()

    # Actual benchmark runs
    times = []
    for _ in range(iterations):
        start = time.perf_counter()
        func()
        end = time.perf_counter()
        times.append(end - start)

    total_time = sum(times)
    avg_time = mean(times)
    std_dev_val = stdev(times) if len(times) > 1 else 0.0

    return BenchmarkResult(
        name=func.__name__ if hasattr(func, '__name__') else str(func),
        iterations=iterations,
        total_time=total_time,
        avg_time=avg_time,
        min_time=min(times),
        max_time=max(times),
        std_dev=std_dev_val,
        ops_per_second=iterations / total_time if total_time > 0 else 0
    )


def profile_function(func: Callable, *args, **kwargs) -> str:
    """Profile a function and return stats as string"""
    profiler = cProfile.Profile()
    profiler.enable()
    func(*args, **kwargs)
    profiler.disable()

    stream = io.StringIO()
    stats = pstats.Stats(profiler, stream=stream)
    stats.sort_stats('cumulative')
    stats.print_stats(20)
    return stream.getvalue()


class TestRT60PerformanceBenchmarks:
    """Performance benchmarks for RT60 calculator"""

    @pytest.fixture
    def rt60_calculator(self):
        """Create RT60 calculator instance"""
        from calculations.enhanced_rt60_calculator import EnhancedRT60Calculator
        return EnhancedRT60Calculator()

    @pytest.fixture
    def sample_space_data_small(self):
        """Small room with 3 surfaces"""
        return {
            'volume': 1200,  # 12x10x10 room
            'surface_instances': [
                {'area': 120, 'material_key': 'gypsum_board_on_studs'},
                {'area': 440, 'material_key': 'painted_concrete_block'},
                {'area': 120, 'material_key': 'carpet_-_heavy_on_concrete'},
            ],
            'target_rt60': 0.6,
            'target_tolerance': 0.1,
            'room_type': 'office'
        }

    @pytest.fixture
    def sample_space_data_medium(self):
        """Medium room with 6 surfaces"""
        return {
            'volume': 15000,
            'surface_instances': [
                {'area': 1500, 'material_key': 'gypsum_board_on_studs'},
                {'area': 1500, 'material_key': 'gypsum_board_on_studs'},
                {'area': 500, 'material_key': 'glass_-_large_panes'},
                {'area': 500, 'material_key': 'painted_concrete_block'},
                {'area': 1500, 'material_key': 'act_nrc_0.70'},
                {'area': 1500, 'material_key': 'carpet_-_heavy_on_concrete'},
            ],
            'target_rt60': 0.8,
            'target_tolerance': 0.1,
            'room_type': 'conference'
        }

    @pytest.fixture
    def sample_space_data_large(self):
        """Large room with 15 surfaces (complex scenario)"""
        surfaces = []
        materials = [
            'gypsum_board_on_studs', 'painted_concrete_block', 'glass_-_large_panes',
            'act_nrc_0.70', 'carpet_-_heavy_on_concrete'
        ]
        for i in range(15):
            surfaces.append({
                'area': 500 + (i * 100),
                'material_key': materials[i % len(materials)]
            })
        return {
            'volume': 100000,
            'surface_instances': surfaces,
            'target_rt60': 1.0,
            'target_tolerance': 0.15,
            'room_type': 'auditorium'
        }

    def test_benchmark_sabine_small_room(self, rt60_calculator, sample_space_data_small):
        """Benchmark Sabine calculation for small room"""
        def run_calc():
            return rt60_calculator.calculate_space_rt60_enhanced(
                sample_space_data_small, method='sabine'
            )

        result = benchmark(run_calc, iterations=100)
        print(f"\n{result}")

        # Performance assertion: should be < 10ms per calculation
        assert result.avg_time < 0.010, f"Sabine (small) too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_sabine_medium_room(self, rt60_calculator, sample_space_data_medium):
        """Benchmark Sabine calculation for medium room"""
        def run_calc():
            return rt60_calculator.calculate_space_rt60_enhanced(
                sample_space_data_medium, method='sabine'
            )

        result = benchmark(run_calc, iterations=100)
        print(f"\n{result}")

        # Performance assertion: should be < 15ms per calculation
        assert result.avg_time < 0.015, f"Sabine (medium) too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_sabine_large_room(self, rt60_calculator, sample_space_data_large):
        """Benchmark Sabine calculation for large room"""
        def run_calc():
            return rt60_calculator.calculate_space_rt60_enhanced(
                sample_space_data_large, method='sabine'
            )

        result = benchmark(run_calc, iterations=100)
        print(f"\n{result}")

        # Performance assertion: should be < 25ms per calculation
        assert result.avg_time < 0.025, f"Sabine (large) too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_eyring_small_room(self, rt60_calculator, sample_space_data_small):
        """Benchmark Eyring calculation for small room"""
        def run_calc():
            return rt60_calculator.calculate_space_rt60_enhanced(
                sample_space_data_small, method='eyring'
            )

        result = benchmark(run_calc, iterations=100)
        print(f"\n{result}")

        # Eyring is more complex, allow slightly more time
        assert result.avg_time < 0.015, f"Eyring (small) too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_eyring_large_room(self, rt60_calculator, sample_space_data_large):
        """Benchmark Eyring calculation for large room"""
        def run_calc():
            return rt60_calculator.calculate_space_rt60_enhanced(
                sample_space_data_large, method='eyring'
            )

        result = benchmark(run_calc, iterations=100)
        print(f"\n{result}")

        # Eyring with large room
        assert result.avg_time < 0.030, f"Eyring (large) too slow: {result.avg_time*1000:.2f}ms"

    def test_profile_rt60_calculation(self, rt60_calculator, sample_space_data_large):
        """Profile RT60 calculation to identify bottlenecks"""
        print("\n--- RT60 Calculation Profile (Top 20 functions) ---")
        profile_output = profile_function(
            rt60_calculator.calculate_space_rt60_enhanced,
            sample_space_data_large,
            method='sabine'
        )
        print(profile_output)


class TestHVACNoisePerformanceBenchmarks:
    """Performance benchmarks for HVAC noise calculation engine"""

    @pytest.fixture
    def hvac_engine(self):
        """Create HVAC noise engine instance"""
        from calculations.hvac_noise_engine import HVACNoiseEngine
        return HVACNoiseEngine()

    @pytest.fixture
    def path_element_class(self):
        """Get PathElement class"""
        from calculations.hvac_noise_engine import PathElement
        return PathElement

    @pytest.fixture
    def simple_path(self, path_element_class):
        """Simple 3-element path"""
        return [
            path_element_class(
                element_type='source',
                element_id='ahu_1',
                source_noise_level=75,
                octave_band_levels=[75, 80, 82, 78, 74, 70, 66, 62]
            ),
            path_element_class(
                element_type='duct',
                element_id='main_duct',
                length=30.0,
                width=24,
                height=18,
                duct_shape='rectangular',
                duct_type='sheet_metal',
                lining_thickness=1.0,
                flow_rate=2000
            ),
            path_element_class(
                element_type='terminal',
                element_id='diffuser_1',
                termination_type='flush'
            )
        ]

    @pytest.fixture
    def medium_path(self, path_element_class):
        """Medium complexity 7-element path"""
        return [
            path_element_class(
                element_type='source',
                element_id='ahu_1',
                source_noise_level=80,
                octave_band_levels=[78, 82, 85, 82, 78, 74, 70, 66]
            ),
            path_element_class(
                element_type='duct',
                element_id='main_supply',
                length=50.0,
                width=30,
                height=24,
                duct_shape='rectangular',
                duct_type='sheet_metal',
                lining_thickness=1.0,
                flow_rate=3000
            ),
            path_element_class(
                element_type='elbow',
                element_id='elbow_1',
                width=30,
                height=24,
                vane_chord_length=4.0,
                num_vanes=3
            ),
            path_element_class(
                element_type='duct',
                element_id='branch_supply',
                length=25.0,
                width=18,
                height=14,
                duct_shape='rectangular',
                duct_type='sheet_metal',
                lining_thickness=1.0,
                flow_rate=1500
            ),
            path_element_class(
                element_type='elbow',
                element_id='elbow_2',
                width=18,
                height=14,
                vane_chord_length=3.0,
                num_vanes=2
            ),
            path_element_class(
                element_type='duct',
                element_id='runout',
                length=10.0,
                width=12,
                height=10,
                duct_shape='rectangular',
                duct_type='sheet_metal',
                lining_thickness=0.5,
                flow_rate=800
            ),
            path_element_class(
                element_type='terminal',
                element_id='diffuser_1',
                termination_type='flush'
            )
        ]

    @pytest.fixture
    def complex_path(self, path_element_class):
        """Complex 15-element path with mixed components"""
        elements = [
            path_element_class(
                element_type='source',
                element_id='ahu_1',
                source_noise_level=85,
                octave_band_levels=[82, 86, 88, 85, 82, 78, 74, 70]
            )
        ]

        # Add multiple duct segments and elbows
        for i in range(6):
            elements.append(path_element_class(
                element_type='duct',
                element_id=f'duct_{i+1}',
                length=20.0 + (i * 5),
                width=24 - (i * 2),
                height=18 - (i),
                duct_shape='rectangular',
                duct_type='sheet_metal',
                lining_thickness=1.0,
                flow_rate=3000 - (i * 300)
            ))
            if i < 5:  # Add elbows between ducts
                elements.append(path_element_class(
                    element_type='elbow',
                    element_id=f'elbow_{i+1}',
                    width=24 - (i * 2),
                    height=18 - i,
                    vane_chord_length=3.0,
                    num_vanes=2
                ))

        # Add silencer
        elements.append(path_element_class(
            element_type='silencer',
            element_id='silencer_1',
            insertion_loss_data={
                '63': 8, '125': 12, '250': 18, '500': 24,
                '1000': 28, '2000': 26, '4000': 20, '8000': 14
            }
        ))

        # Add terminal
        elements.append(path_element_class(
            element_type='terminal',
            element_id='diffuser_1',
            termination_type='flush'
        ))

        return elements

    @pytest.fixture
    def receiver_config(self):
        """Standard receiver room configuration"""
        return {
            'room_volume': 12000,
            'room_absorption': 400,
            'distance_from_terminal': 8.0,
            'termination_type': 'flush'
        }

    def test_benchmark_simple_path(self, hvac_engine, simple_path, receiver_config):
        """Benchmark simple path calculation"""
        def run_calc():
            return hvac_engine.calculate_path_noise(simple_path, receiver_config)

        result = benchmark(run_calc, iterations=50)
        print(f"\n{result}")

        # Performance assertion: should be < 50ms per calculation
        assert result.avg_time < 0.050, f"Simple path too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_medium_path(self, hvac_engine, medium_path, receiver_config):
        """Benchmark medium path calculation"""
        def run_calc():
            return hvac_engine.calculate_path_noise(medium_path, receiver_config)

        result = benchmark(run_calc, iterations=50)
        print(f"\n{result}")

        # Performance assertion: should be < 100ms per calculation
        assert result.avg_time < 0.100, f"Medium path too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_complex_path(self, hvac_engine, complex_path, receiver_config):
        """Benchmark complex path calculation"""
        def run_calc():
            return hvac_engine.calculate_path_noise(complex_path, receiver_config)

        result = benchmark(run_calc, iterations=50)
        print(f"\n{result}")

        # Performance assertion: should be < 200ms per calculation
        assert result.avg_time < 0.200, f"Complex path too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_nc_rating_calculation(self, hvac_engine):
        """Benchmark NC rating calculation"""
        from calculations.nc_rating_analyzer import NCRatingAnalyzer, OctaveBandData

        analyzer = NCRatingAnalyzer()
        octave_data = OctaveBandData(
            freq_63=55, freq_125=50, freq_250=45, freq_500=40,
            freq_1000=35, freq_2000=32, freq_4000=30, freq_8000=28
        )

        def run_calc():
            return analyzer.analyze_octave_band_data(octave_data)

        result = benchmark(run_calc, iterations=200)
        print(f"\n{result}")

        # NC rating should be very fast
        assert result.avg_time < 0.005, f"NC rating too slow: {result.avg_time*1000:.2f}ms"

    def test_profile_hvac_calculation(self, hvac_engine, complex_path, receiver_config):
        """Profile HVAC calculation to identify bottlenecks"""
        print("\n--- HVAC Path Calculation Profile (Top 20 functions) ---")
        profile_output = profile_function(
            hvac_engine.calculate_path_noise,
            complex_path,
            receiver_config
        )
        print(profile_output)


class TestDatabasePerformanceBenchmarks:
    """Performance benchmarks for database operations"""

    @pytest.fixture
    def temp_database(self, tmp_path):
        """Create temporary database for testing"""
        import os
        os.environ['DATABASE_PATH'] = str(tmp_path / 'test_perf.db')

        from models import initialize_database, get_session, Project, Space, close_database
        initialize_database()

        yield get_session, Project, Space

        close_database()

    def test_benchmark_project_creation(self, temp_database):
        """Benchmark project creation"""
        get_session, Project, Space = temp_database

        def run_creation():
            session = get_session()
            project = Project(
                name=f"Test Project",
                description="Performance test project",
                default_scale="1:100",
                default_units="feet"
            )
            session.add(project)
            session.commit()
            project_id = project.id
            session.close()
            return project_id

        result = benchmark(run_creation, iterations=50)
        print(f"\n{result}")

        # Database write should be < 20ms
        assert result.avg_time < 0.020, f"Project creation too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_project_query(self, temp_database):
        """Benchmark project query"""
        get_session, Project, Space = temp_database

        # Create test data
        session = get_session()
        for i in range(100):
            project = Project(
                name=f"Query Test Project {i}",
                description=f"Test project {i}",
            )
            session.add(project)
        session.commit()
        session.close()

        def run_query():
            session = get_session()
            projects = session.query(Project).filter(
                Project.name.like('%Query Test%')
            ).all()
            count = len(projects)
            session.close()
            return count

        result = benchmark(run_query, iterations=100)
        print(f"\n{result}")

        # Query should be < 10ms
        assert result.avg_time < 0.010, f"Project query too slow: {result.avg_time*1000:.2f}ms"


class TestMaterialSearchPerformanceBenchmarks:
    """Performance benchmarks for material search operations"""

    @pytest.fixture
    def material_search(self):
        """Create material search instance"""
        try:
            from data.materials import MaterialSearch
            return MaterialSearch()
        except ImportError:
            pytest.skip("MaterialSearch not available")

    def test_benchmark_material_search_by_category(self, material_search):
        """Benchmark material search by category"""
        def run_search():
            return material_search.search_by_category('ceiling')

        result = benchmark(run_search, iterations=100)
        print(f"\n{result}")

        # Search should be < 10ms
        assert result.avg_time < 0.010, f"Category search too slow: {result.avg_time*1000:.2f}ms"

    def test_benchmark_material_search_by_nrc(self, material_search):
        """Benchmark material search by NRC range"""
        def run_search():
            return material_search.search_by_nrc_range(0.6, 0.9)

        result = benchmark(run_search, iterations=100)
        print(f"\n{result}")

        # Search should be < 15ms
        assert result.avg_time < 0.015, f"NRC search too slow: {result.avg_time*1000:.2f}ms"


class TestExcelExportPerformanceBenchmarks:
    """Performance benchmarks for Excel export operations"""

    def test_benchmark_excel_workbook_creation(self, tmp_path):
        """Benchmark Excel workbook creation with openpyxl directly"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        output_path = tmp_path / "perf_test.xlsx"

        def run_export():
            wb = Workbook()
            ws = wb.active
            ws.title = "Performance Test"

            # Header row
            headers = ['Room', 'Volume', 'RT60', 'Target', 'Compliance', 'NC Rating']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            # Data rows (simulate 50 rooms)
            for i in range(50):
                ws.cell(row=i+2, column=1, value=f'Room {i+1}')
                ws.cell(row=i+2, column=2, value=1000 + i*100)
                ws.cell(row=i+2, column=3, value=round(0.5 + i*0.02, 2))
                ws.cell(row=i+2, column=4, value=0.6)
                ws.cell(row=i+2, column=5, value='Pass' if i % 2 == 0 else 'Fail')
                ws.cell(row=i+2, column=6, value=35 + (i % 10))

            # Add a second sheet for HVAC data
            ws2 = wb.create_sheet("HVAC Paths")
            hvac_headers = ['Path', 'Source', 'Terminal', 'NC', 'dBA']
            for col, header in enumerate(hvac_headers, 1):
                ws2.cell(row=1, column=col, value=header)

            for i in range(20):
                ws2.cell(row=i+2, column=1, value=f'Path {i+1}')
                ws2.cell(row=i+2, column=2, value=f'AHU-{i+1}')
                ws2.cell(row=i+2, column=3, value=f'Diffuser-{i+1}')
                ws2.cell(row=i+2, column=4, value=35 + i)
                ws2.cell(row=i+2, column=5, value=40 + i)

            wb.save(str(output_path))

        result = benchmark(run_export, iterations=20)
        print(f"\n{result}")

        # Excel creation should be < 100ms
        assert result.avg_time < 0.100, f"Excel creation too slow: {result.avg_time*1000:.2f}ms"


class TestPerformanceSummary:
    """Generate overall performance summary"""

    def test_generate_performance_report(self):
        """Generate a performance summary report"""
        print("\n" + "="*60)
        print("PERFORMANCE PROFILING SUMMARY")
        print("="*60)
        print("""
Performance Targets:
- RT60 Calculation (small room): < 10ms
- RT60 Calculation (large room): < 25ms
- HVAC Path (simple): < 50ms
- HVAC Path (complex): < 200ms
- NC Rating Calculation: < 5ms
- Database Operations: < 20ms
- Material Search: < 15ms
- Excel Export: < 500ms

Run individual benchmark tests with -v -s flags to see detailed timing:
  pytest tests/test_performance_profiling.py -v -s

To profile a specific calculation and identify bottlenecks:
  pytest tests/test_performance_profiling.py::TestRT60PerformanceBenchmarks::test_profile_rt60_calculation -v -s
  pytest tests/test_performance_profiling.py::TestHVACNoisePerformanceBenchmarks::test_profile_hvac_calculation -v -s
        """)
        print("="*60)


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
